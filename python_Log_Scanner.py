#!/usr/bin/env python
import os
import glob
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import shutil
from datetime import date, timedelta
import time
from openpyxl import load_workbook
import xlrd
from scannerlib import read_atti_from_property_file,filter_file,read_atti_from_err_patternfile,create_folder
from scannerlib import export_excel,export_excel_for_today,export_text
from scannerlib import mail,implementChart

# Main Program
if __name__ == '__main__':

    start = time.time()
    Check_list = 0
    support_file_path =  r"/new/server/sphere/v48/appserver1/profiles12/Python_Logs_Scanner_Script"   # script location

    Propertydic = read_atti_from_property_file(support_file_path)

    word_list_to_remove = ['] INFO[',']DEBUG[']  # removing info and debug logs from file

    log_date = int(Propertydic['dateLog'])

    today = datetime.now() - timedelta(log_date)

    listOfDate, Errordata = [], []

    countofcriticalrows = 0

    if Propertydic['value'] == "True":
        try:
            Input_file_paths = [Propertydic['path_of_log_file'], Propertydic['path_of_log_file1'],
                                Propertydic['path_of_log_file2'],
                                Propertydic['path_of_log_file3']]
            print("Input_file_paths",Input_file_paths)
        except:
            print("Error : Input_path_of_file are wrong in ProgramConfig.property file")
        for input_file in Input_file_paths:

            if input_file == "NA":
                print("Please insert the file Path Correctly ")
            else:
                List_of_files = glob.glob(input_file + Propertydic['fileLocation'] + Propertydic['Scanfile'])
                #print("List_of_files",List_of_files)
                dir_contents = os.listdir('.')
                if dir_contents:
                    try:
                        Sorted_list_of_file = filter_file(List_of_files)
                        # Create folder 
                        folder_path = str(
                            create_folder(input_file + "//Logs_Scanner_logs//Log_" + today.strftime('%d_%m_%Y')))  

                        folder_name = "Log_" + today.strftime('%d_%m_%Y')
                        print("#", folder_name, "Folder is created")
                    except OSError:
                        print("Error : Creating Folder")
                    save_text_file = folder_path + "//ErrorsAdvantageApps.txt"
                    save_excel_file = folder_path + "//ErrorsAdvantageApps.xlsx"
                    save_excel_file_today = folder_path + "//ErrorsAdvantageAppsfortoday.xlsx"
                    if Propertydic['export_text'] == "True":
                        export_text(Sorted_list_of_file,save_text_file,word_list_to_remove)  # Calling read write function for text
                        print("Chunking Started .......")
                        fs = FileSplit(file=save_text_file, splitsize=20000000,output_dir=folder_path)
                        fs.split(include_header=True)
                        os.remove(save_text_file)
                    else:
                        print("export_text config :", Propertydic['export_text'])
                    if Propertydic['export_excel'] == "True":
                        export_excel(Sorted_list_of_file,save_excel_file)  # Calling read write function for excel
                    else:
                        print("export_excel config :", Propertydic['export_excel'])
                    Check_list = 2
                    if Propertydic['export_excel_for_today'] == "True":
                        countofcriticalrows = export_excel_for_today(Sorted_list_of_file,save_excel_file_today,today)  # Calling read write function for excel today
                        
                        print("critical_logs_count:",countofcriticalrows )

                        for i in range(15):
                            dt = date.today() - timedelta(i + 1)
                            try:
                                listOfDate.append(dt.strftime('%d_%m_%Y'))
                                book = xlrd.open_workbook(
                                    Propertydic['UnifyAttachedfile'] + "//Logs_Scanner_logs//Log_" + dt.strftime('%d_%m_%Y') + "//ErrorsAdvantageAppsfortoday.xlsx")
                                sheet = book.sheet_by_index(0)
                                
                                Errordata.append(sheet.nrows - 1)
                            except:
                                print("Folder of this date not exist", dt.strftime('%d_%m_%Y'))
                                Errordata.append("NA")
                        implementChart(listOfDate, Errordata, save_excel_file_today)
                    else:
                        print("export_excel_for_today config :", Propertydic['export_excel_for_today'])
                    countofrows = 0
                    Check_list = Check_list + 1
                    print("# Three ErrorsAdvantageApps files are generated in", folder_name, "folder")

                    for i in range(15):
                        dt = date.today() - timedelta(i + 48)
                        today = date.today()
                        value = today.strftime('%d_%m_%Y')
                        delete_file = dt.strftime('%d_%m_%Y')
                        try:
                            shutil.rmtree(input_file + "//Unify_Logs_Scanner_logs//Log_" + delete_file)
                            print("Folder is deleted", delete_file)
                        except:
                            print("Folder ", delete_file, " not available")
                else:
                    print("Scan file are not present in the location")
        try:
            today = datetime.now() - timedelta(log_date)
            #print("today:",today)
            wb = load_workbook(Propertydic['Attachedfile'] + "//Logs_Scanner_logs//Log_" + today.strftime(
                '%d_%m_%Y') + "//ErrorsAdvantageAppsfortoday.xlsx", read_only=True)
            sheet = wb.worksheets[0]
            countofrows = sheet.max_row
            print("countofrows", countofrows)
        except:
            print("Error : No such file or directory:",
                  Propertydic['Attachedfile'] + "/Logs_Scanner_logs/Log_" + today.strftime(
                      '%d_%m_%Y') + "/ErrorsAdvantageAppsfortoday.xlsx")

        if Propertydic['SendMail'] == "True" and Propertydic['export_excel_for_today'] == "True":
            if countofrows == 1:
                Email_Body = Propertydic['Email_Body_Excelfile_Empty']
                fromaddr = Propertydic['Send_name']
                toaddr = Propertydic['Receiver_name']
                msg = MIMEMultipart()
                msg['From'] = fromaddr
                msg['To'] = toaddr
                msg['Subject'] = (Propertydic['Email_Subject'] + today.strftime('%d_%m_%Y'))
                body = Email_Body
                msg.attach(MIMEText(body, 'plain'))
                text = msg.as_string()
            elif Check_list == 3:
                #print("Check_list", Check_list)
                Email_Body = Propertydic['Email_Body_success'] + "\n Total number of errors : " + str(
                    countofrows - 1) + "\n Count of critical errors : " + str(countofcriticalrows)
                fromaddr = Propertydic['Send_name']
                toaddr = Propertydic['Receiver_name']
                msg = MIMEMultipart()
                msg['From'] = fromaddr
                msg['To'] = toaddr
                msg['Subject'] = (Propertydic['Email_Subject'] + today.strftime('%d_%m_%Y'))
                body = Email_Body
                msg.attach(MIMEText(body, 'plain'))
                filename = 'ErrorsAdvantageAppsfortoday.xlsx'
                attachment = open(Propertydic['UnifyAttachedfile'] + "/Logs_Scanner_logs/Log_" + today.strftime(
                    '%d_%m_%Y') + "/ErrorsAdvantageAppsfortoday.xlsx", 'rb')
                #print(attachment)
                p = MIMEBase('application', 'octet-stream')
                p.set_payload((attachment).read())
                encoders.encode_base64(p)
                p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
                msg.attach(p)
                text = msg.as_string()
            else:
                Email_Body = Propertydic['Email_Body_failed']
                fromaddr = Propertydic['Send_name']
                toaddr = Propertydic['Receiver_name']
                msg = MIMEMultipart()
                msg['From'] = fromaddr
                msg['To'] = toaddr
                msg['Subject'] = (Propertydic['Email_Subject'] + today.strftime('%d_%m_%Y'))
                body = Email_Body
                msg.attach(MIMEText(body, 'plain'))
                text = msg.as_string()
            mail(Email_Body,toaddr,fromaddr,text)
        else:
            print("Mail configuration is : ", Propertydic['SendMail'])
    end = time.time()
    print("Total time taken to execute is ", end - start)
