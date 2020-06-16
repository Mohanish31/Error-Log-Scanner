#!/usr/bin/env python
import re
import os
import xlsxwriter
import smtplib
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import PieChart, Reference

# Read all the value from Parameter file
def read_atti_from_property_file(support_file_path):
    separator = "="
    keys = {}
    try:
        with open(support_file_path + '/ParamConfig.properties') as f:
            for line in f:
                if separator in line:
                    name, value = line.split(separator, 1)
                    keys[name.strip()] = value.strip()
            print("Completed : Dictionary Successfully Created of Property File ")
    except IOError:
        print('Error : Property file is not avaiable')
    return keys


def read_atti_from_err_patternfile(support_file_path):
    separator = "="
    keys = {}
    try:
        with open(support_file_path + '/Error_Pattern.txt') as parameter_Err_patternfile:
            for line in parameter_Err_patternfile:
                if separator in line:
                    name, value = line.split(separator, 1)
                    keys[name.strip()] = value.strip()
            print("Completed : Dictionary Successfully Created of Error_Pattern_file ")
    except IOError:
        print('Error : Error Pattern file is not avaiable')
    return keys


# Create a folder on todays day

def create_folder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print('Error: Creating directory. ' + directory)
    return directory


# Sort the alphanumic list of log files

def sorted_file_alphanumic(list_of_final_files):
    convert = lambda text: int(text) if text.isdigit() else text
    alphanum_key = lambda key: [convert(c) for c in re.split('([0-9]+)', key)]
    return sorted(list_of_final_files, key=alphanum_key)


# Remove unnessary file from the sorted list

def filter_file(list_of_files):
    list_of_final_files = []
    for list1 in list_of_files:
        list_word = ['TargetGroupsError.log', 'TargetGroupsError-1.log', 'TargetGroupsError-2.log',
                     'TargetGroupsError-3.log', 'TargetGroupsError-4.log', 'TargetGroupsError-5.log',
                     'AdvantageAppsError.log', 'AdvantageAppsError-1.log', 'AdvantageAppsError-2.log',
                     'AdvantageAppsError-3.log', 'AdvantageAppsError-4.log', 'AdvantageAppsError-5.log',
                     'AdvantageAppsError-6.log', 'AdvantageAppsError-7.log', 'AdvantageAppsError-8.log',
                     'AdvantageAppsError-9.log', 'AdvantageAppsError-10.log'
                     ]
        if not any(list_word in list1 for list_word in list_word):
            list_of_final_files.append(list1)
    Sorted_list_of_final_files = sorted_file_alphanumic(list_of_final_files)
    Sorted1_list_of_final_files = Sorted_list_of_final_files[-1:] + Sorted_list_of_final_files[:-1]
    return Sorted1_list_of_final_files


# simplify a error log line.

def extract_data(value):
    try:
        new_value = value.replace(']', '|', 5)
        new_value = new_value.replace(',', '|', 1)
        new_value = new_value.replace("[", '|', 2).replace('|', '[', 1)
        new_value = new_value.replace("[", '', 4)
        return new_value
    except:
        print("Error : Extract_data Function issue")


# Generate a text file which contain all the line

def export_text(Sorted_list_of_file,save_text_file,word_lists):

    for file_name in Sorted_list_of_file:
        print("Current File Being Processed for text is: ", file_name)
        with open(file_name) as f:
            with open(save_text_file, "a+") as f1:
                for line in f:
                    if not any(word_list in line for word_list in word_lists):
                        f1.write(line)
# Generate a excel file which contain all error the line

def export_excel(Sorted_list_of_file,save_excel_file):
    workbook = xlsxwriter.Workbook(save_excel_file)
    worksheet = workbook.add_worksheet()
    critical_word_lists = ['CRITICAL', 'java.lang.NullPointerException', 'FATAL']
    wordlog_lists = [']ERROR[', ']FATAL[']
    countofcriticalrows = 0
    row = 1  # row counter
    worksheet.write(0, 0, "Date/Time")
    worksheet.write(0, 1, "Thread Count")
    worksheet.write(0, 2, "Logs")
    worksheet.write(0, 3, "WebContainer/Thread/server.startup")
    worksheet.write(0, 4, "Session ID")
    worksheet.write(0, 5, "ID")
    worksheet.write(0, 6, "Path")
    worksheet.write(0, 7, "Details")
    for file_name in Sorted_list_of_file:
        print("Current File Being Processed for all_Excel is : ", file_name)
        with open(file_name) as f:
            for line in f:
                # print(line)
                if any(wordlog_list in line for wordlog_list in wordlog_lists):
                    # print(wordlog_lists)
                    line = extract_data(line)
                    line = line.split("|")
                    col = 0
                    for line_value in line:
                        worksheet.write(row, col, line_value)
                        # print(row, col, line_value)
                        if any(critical_word_list in line_value for critical_word_list in critical_word_lists):
                            countofcriticalrows = countofcriticalrows + 1
                            # print("critical_word_lists")
                        col += 1
                    row += 1
        countofcriticalrows = countofcriticalrows
    countofcriticalrows = countofcriticalrows
    workbook.close()


# Generate a excel file which contain all error the line only for today

def export_excel_for_today(Sorted_list_of_file,save_excel_file_today,today):
    workbook = xlsxwriter.Workbook(save_excel_file_today)
    worksheet = workbook.add_worksheet()
    day = today.strftime('%Y-%m-%d')
    # print(day)
    row = 1  # row counter in file
    countofcriticalrows = 0
    critical_word_lists = ['CRITICAL', 'java.lang.NullPointerException', 'FATAL']
    wordlog_lists = [']ERROR[', ']FATAL[']
    worksheet.write(0, 0, "Date/Time")
    worksheet.write(0, 1, "Thread Count")
    worksheet.write(0, 2, "Logs")
    worksheet.write(0, 3, "WebContainer/Thread/server.startup")
    worksheet.write(0, 4, "Session ID")
    worksheet.write(0, 5, "ID")
    worksheet.write(0, 6, "Path")
    worksheet.write(0, 7, "Details")
    for file_name in Sorted_list_of_file:
        print("Current File Being Processed for Excel is : ", file_name)
        with open(file_name) as f:
            for line in f:
                if day in line:
                    if any(wordlog_list in line for wordlog_list in wordlog_lists):
                        # print(wordlog_lists)
                        line = extract_data(line)
                        line = line.split("|")
                        col = 0
                        for line_value in line:
                            worksheet.write(row, col, line_value)
                            # print(row, col, line_value)
                            if any(critical_word_list in line_value for critical_word_list in critical_word_lists):
                                countofcriticalrows = countofcriticalrows + 1
                                # print("critical_word_lists")
                            col += 1
                        row += 1
        countofcriticalrows = countofcriticalrows
    countofcriticalrows = countofcriticalrows
    workbook.close()
    return countofcriticalrows

def mail(Email_Body,toaddr,fromaddr,text):
    s = smtplib.SMTP('smtpna.infores.com')

    # sending the mail
    for reception in toaddr.split(','):
        s.sendmail(fromaddr, reception, text)
        print(reception)
        print("sent mail")
        # terminating the session
    s.quit()


def implementChart(listOfDate, Errordate, save_excel_file_today):
    wb_obj = openpyxl.load_workbook(save_excel_file_today)
    wb_obj.create_sheet('Chart view')
    sheet_obj = wb_obj["Chart view"]
    Drows = 1
    Erows = 1
    cell_obj = sheet_obj.cell(row=1, column=1)
    cell_obj.value = "listOfDate"
    cell_obj = sheet_obj.cell(row=1, column=2)
    cell_obj.value = "Error_data"

    for i in listOfDate:
        cell_obj = sheet_obj.cell(row=Drows + 1, column=1)
        cell_obj.value = i
        Drows += 1
    for i in Errordate:
        cell_obj = sheet_obj.cell(row=Erows + 1, column=2)
        cell_obj.value = i
        Erows += 1

    dates = Reference(sheet_obj, min_col=1, min_row=2, max_col=1, max_row=16)
    values = Reference(sheet_obj, min_col=2, min_row=2, max_col=2, max_row=16)

    chart = BarChart()
    chart.add_data(values)
    chart.set_categories(dates)
    chart.height = 15  # default is 7.5
    chart.width = 30
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.title = " 15 Day's Error Count "
    chart.x_axis.title = " Dates "
    chart.y_axis.title = " Number of Error "
    sheet_obj.add_chart(chart, "A1")
    wb_obj.save(save_excel_file_today)
